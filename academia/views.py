import csv
import io

from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.models import User, Group
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Curso, Material, Inscripcion, Calificacion, LogActividad, PeriodoAcademico,BannerCarrusel, ConfiguracionLanding,Curso, Examen, Pregunta, Opcion,IntentoExamen,RespuestaEstudiante
from .forms import RegistroUsuarioForm, EditarUsuarioForm, CursoForm, InscripcionForm, PreguntaForm


# ----------------------------------------------------
# UTILIDAD DE AUDITORÍA (LOGS)
# ----------------------------------------------------
def registrar_log(request, accion, detalles=""):
    """Registra una acción en la tabla de auditoría con la IP del usuario."""
    ip = request.META.get('HTTP_X_FORWARDED_FOR')
    if ip:
        ip = ip.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')

    LogActividad.objects.create(
        usuario=request.user if request.user.is_authenticated else None,
        accion=accion,
        detalles=detalles,
        ip_origen=ip
    )


# ----------------------------------------------------
# VALIDACIÓN DE ACCESOS
# ----------------------------------------------------
def es_administrador(user):
    return user.is_authenticated and (user.is_superuser or user.is_staff or user.groups.filter(name='Administrador').exists())


def es_docente_valido(user):
    return user.is_authenticated and (user.groups.filter(name='Docentes').exists() or user.is_staff or user.is_superuser)


# ----------------------------------------------------
# VISTAS PÚBLICAS Y GENERALES
# ----------------------------------------------------
def inicio_publico(request):
    periodo_activo = PeriodoAcademico.objects.filter(activo=True).first()
    
    # Solo precargar 'periodo', que es la ForeignKey válida en Curso
    cursos_qs = Curso.objects.select_related('periodo')

    if periodo_activo:
        cursos = cursos_qs.filter(periodo=periodo_activo)
        if not cursos.exists():
            cursos = cursos_qs.all()
    else:
        cursos = cursos_qs.all()

    banners = BannerCarrusel.objects.filter(activo=True).order_by('orden')
    config_landing, _ = ConfiguracionLanding.objects.get_or_create(id=1)

    context = {
        'periodo_activo': periodo_activo,
        'cursos_destacados': cursos,
        'banners': banners,
        'config': config_landing,
    }
    return render(request, 'inicio_publico.html', context)

def salir(request):
    logout(request)
    return redirect('/cuentas/login/')


# ----------------------------------------------------
# CONTROLADOR PRINCIPAL SEGÚN ROL
# ----------------------------------------------------
@login_required 
def dashboard(request):
    # 1. Administrador -> Panel de administración
    if es_administrador(request.user):
        return redirect('admin_dashboard')
    
    # 2. Docente -> Panel de gestión docente
    if request.user.groups.filter(name='Docentes').exists():
        return redirect('panel_docente')
    
    # 3. Alumno -> Dashboard con sus cursos
    mis_inscripciones = Inscripcion.objects.filter(alumno=request.user).select_related('curso')
    context = {
        'inscripciones': mis_inscripciones,
        'es_docente': False
    }
    return render(request, 'dashboard.html', context)


# ----------------------------------------------------
# SECCIÓN ADMINISTRACIÓN: USUARIOS
# ----------------------------------------------------
@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_dashboard(request):
    total_alumnos = User.objects.filter(groups__name='Alumnos').count()
    total_docentes = User.objects.filter(groups__name='Docentes').count()
    total_cursos = Curso.objects.count()
    
    # Obtener el período académico activo para el modal de gestión
    periodo_activo = PeriodoAcademico.objects.filter(activo=True).first()
    
    query = request.GET.get('q', '').strip()
    rol_filtro = request.GET.get('rol', '').strip()
    
    usuarios_qs = User.objects.all().prefetch_related('groups').order_by('-date_joined')

    if query:
        usuarios_qs = usuarios_qs.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )

    if rol_filtro == 'admin':
        usuarios_qs = usuarios_qs.filter(Q(is_superuser=True) | Q(is_staff=True))
    elif rol_filtro == 'docente':
        usuarios_qs = usuarios_qs.filter(groups__name='Docentes')
    elif rol_filtro == 'alumno':
        usuarios_qs = usuarios_qs.filter(groups__name='Alumnos')

    paginator = Paginator(usuarios_qs, 15)
    page_number = request.GET.get('page')
    usuarios = paginator.get_page(page_number)

    context = {
        'total_alumnos': total_alumnos,
        'total_docentes': total_docentes,
        'total_cursos': total_cursos,
        'usuarios': usuarios,
        'query': query,
        'rol_filtro': rol_filtro,
        'periodo_activo': periodo_activo,
    }
    return render(request, 'admin_dashboard.html', context)

@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_detalle_usuario(request, user_id):
    usuario_detalle = get_object_or_404(User, id=user_id)
    
    # Cursos matriculados como alumno
    inscripciones = Inscripcion.objects.filter(alumno=usuario_detalle).select_related('curso')
    
    # Cursos asignados como docente (relación ManyToMany con Curso)
    cursos_docente = Curso.objects.filter(docentes=usuario_detalle)
    
    context = {
        'usuario_detalle': usuario_detalle,
        'inscripciones': inscripciones,
        'cursos_docente': cursos_docente,
    }
    return render(request, 'admin_detalle_usuario.html', context)


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def registrar_usuario(request):
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            nuevo_usuario = form.save()
            rol_nombre = form.cleaned_data.get('rol', 'Sin rol')
            registrar_log(request, "Creación de Usuario", f"Creó al usuario '{nuevo_usuario.username}' con rol '{rol_nombre}'")
            messages.success(request, f"Usuario '{nuevo_usuario.username}' registrado correctamente.")
            return redirect('admin_dashboard')
    else:
        form = RegistroUsuarioForm()

    return render(request, 'registro_usuario.html', {'form': form})


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def editar_usuario(request, user_id):
    usuario_editar = get_object_or_404(User, id=user_id)
    rol_actual = 'Alumno'
    if usuario_editar.is_superuser or usuario_editar.is_staff:
        rol_actual = 'Administrador'
    elif usuario_editar.groups.filter(name='Docentes').exists():
        rol_actual = 'Docente'

    if request.method == 'POST':
        usuario_editar.username = request.POST.get('username')
        usuario_editar.first_name = request.POST.get('first_name')
        usuario_editar.last_name = request.POST.get('last_name')
        usuario_editar.email = request.POST.get('email')

        nuevo_password = request.POST.get('password')
        if nuevo_password:
            usuario_editar.set_password(nuevo_password)

        nuevo_rol = request.POST.get('rol')
        usuario_editar.groups.clear()
        if nuevo_rol == 'Administrador':
            usuario_editar.is_staff = True
            usuario_editar.is_superuser = True
        elif nuevo_rol == 'Docente':
            usuario_editar.is_staff = False
            usuario_editar.is_superuser = False
            grupo_docentes, _ = Group.objects.get_or_create(name='Docentes')
            usuario_editar.groups.add(grupo_docentes)
        else:  # Alumno
            usuario_editar.is_staff = False
            usuario_editar.is_superuser = False
            grupo_alumnos, _ = Group.objects.get_or_create(name='Alumnos')
            usuario_editar.groups.add(grupo_alumnos)

        usuario_editar.save()
        messages.success(request, f"Usuario @{usuario_editar.username} actualizado correctamente.")
        return redirect('admin_dashboard')

    context = {
        'usuario_editar': usuario_editar,
        'rol_actual': rol_actual,
    }
    return render(request, 'editar_usuario.html', context)


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_matricular(request, curso_id=None):
    cursos = Curso.objects.filter(estado=True).order_by('titulo')
    curso_seleccionado = None
    alumnos_matriculados_ids = []
    
    grupo_alumnos = Group.objects.filter(name='Alumnos').first()
    alumnos = User.objects.filter(groups=grupo_alumnos).order_by('last_name', 'first_name') if grupo_alumnos else User.objects.none()

    if curso_id:
        curso_seleccionado = get_object_or_404(Curso, id=curso_id)
        alumnos_matriculados_ids = list(Inscripcion.objects.filter(curso=curso_seleccionado).values_list('alumno_id', flat=True))

    if request.method == 'POST':
        curso_post_id = request.POST.get('curso_id')
        curso_actual = get_object_or_404(Curso, id=curso_post_id)
        seleccionados_ids = request.POST.getlist('alumnos_seleccionados')
        seleccionados_ids = [int(i) for i in seleccionados_ids]

        Inscripcion.objects.filter(curso=curso_actual).exclude(alumno_id__in=seleccionados_ids).delete()

        nuevos_matriculados = 0
        for a_id in seleccionados_ids:
            obj, created = Inscripcion.objects.get_or_create(curso=curso_actual, alumno_id=a_id)
            if created:
                nuevos_matriculados += 1

        LogActividad.objects.create(
            usuario=request.user,
            accion=f"Actualizó matrícula del curso '{curso_actual.titulo}' ({len(seleccionados_ids)} alumnos activos)"
        )

        messages.success(request, f"Matrícula actualizada exitosamente para '{curso_actual.titulo}'.")
        return redirect('admin_matricular_curso', curso_id=curso_actual.id)

    context = {
        'cursos': cursos,
        'curso_seleccionado': curso_seleccionado,
        'alumnos': alumnos,
        'alumnos_matriculados_ids': alumnos_matriculados_ids,
    }
    return render(request, 'admin_matricular.html', context)


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def eliminar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    
    if usuario == request.user:
        messages.error(request, "No puedes eliminar tu propia cuenta de administrador.")
        return redirect('admin_dashboard')
    
    nombre = usuario.username
    registrar_log(request, "Eliminación de Usuario", f"Eliminó la cuenta del usuario '{nombre}'")
    usuario.delete()
    messages.success(request, f"El usuario '{nombre}' fue eliminado correctamente.")
    return redirect('admin_dashboard')


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def eliminar_usuarios_masivo(request):
    if request.method == 'POST':
        user_ids = request.POST.getlist('usuarios_seleccionados')
        
        if not user_ids:
            messages.error(request, "No seleccionaste ningún usuario para eliminar.")
            return redirect('admin_dashboard')
        
        usuarios_a_borrar = User.objects.filter(id__in=user_ids).exclude(id=request.user.id)
        cantidad = usuarios_a_borrar.count()
        nombres = ", ".join(usuarios_a_borrar.values_list('username', flat=True))
        usuarios_a_borrar.delete()
        
        registrar_log(request, "Eliminación Masiva", f"Eliminó {cantidad} usuario(s): {nombres}")
        messages.success(request, f"Se eliminaron {cantidad} usuario(s) correctamente.")
    
    return redirect('admin_dashboard')


# ----------------------------------------------------
# SECCIÓN ADMINISTRACIÓN: CURSOS Y MATRÍCULAS
# ----------------------------------------------------
@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_cursos_lista(request):
    cursos = Curso.objects.prefetch_related('docentes').all()
    
    total_alumnos = User.objects.filter(groups__name='Alumnos').count()
    total_docentes = User.objects.filter(groups__name='Docentes').count()
    total_cursos = Curso.objects.filter(estado=True).count()

    context = {
        'cursos': cursos,
        'total_alumnos': total_alumnos,
        'total_docentes': total_docentes,
        'total_cursos': total_cursos,
    }
    return render(request, 'admin_cursos_lista.html', context)


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_crear_curso(request):
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        descripcion = request.POST.get('descripcion')
        estado = True if request.POST.get('estado') else False
        imagen = request.FILES.get('imagen_portada')
        
        curso = Curso.objects.create(
            titulo=titulo,
            descripcion=descripcion,
            estado=estado,
            imagen_portada=imagen
        )
        
        docentes_ids = request.POST.getlist('docentes')
        if docentes_ids:
            curso.docentes.set(docentes_ids)
            
        messages.success(request, f"Curso '{curso.titulo}' creado exitosamente.")
        return redirect('admin_cursos_lista')

    docentes = User.objects.filter(groups__name='Docentes')
    return render(request, 'admin_crear_curso.html', {'docentes': docentes})


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_editar_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)
    
    if request.method == 'POST':
        curso.titulo = request.POST.get('titulo')
        curso.descripcion = request.POST.get('descripcion')
        curso.estado = True if request.POST.get('estado') else False
        
        if 'imagen_portada' in request.FILES:
            curso.imagen_portada = request.FILES['imagen_portada']
            
        docentes_ids = request.POST.getlist('docentes')
        curso.docentes.set(docentes_ids)
        curso.save()
        
        messages.success(request, f"Curso '{curso.titulo}' actualizado.")
        return redirect('admin_cursos_lista')

    docentes = User.objects.filter(groups__name='Docentes')
    return render(request, 'editar_curso.html', {'curso': curso, 'docentes': docentes})


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_eliminar_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)
    titulo = curso.titulo
    registrar_log(request, "Eliminación de Curso", f"Eliminó el curso '{titulo}'")
    curso.delete()
    messages.success(request, f"El curso '{titulo}' fue eliminado correctamente.")
    return redirect('admin_cursos_lista')


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_matricular_alumno(request):
    if request.method == 'POST':
        form = InscripcionForm(request.POST)
        if form.is_valid():
            inscripcion = form.save()
            registrar_log(request, "Matrícula de Alumno", f"Matriculó a '{inscripcion.alumno.username}' en '{inscripcion.curso.titulo}'")
            messages.success(request, f"Alumno '{inscripcion.alumno.username}' matriculado en '{inscripcion.curso.titulo}'.")
            return redirect('admin_cursos_lista')
    else:
        form = InscripcionForm()
    return render(request, 'admin_matricular.html', {'form': form})


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_curso_alumnos(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)
    inscripciones = Inscripcion.objects.filter(curso=curso).select_related('alumno').order_by('alumno__last_name', 'alumno__first_name')
    
    context = {
        'curso': curso,
        'inscripciones': inscripciones,
    }
    return render(request, 'admin_curso_alumnos.html', context)


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_desmatricular_alumno(request, inscripcion_id):
    inscripcion = get_object_or_404(Inscripcion, id=inscripcion_id)
    curso_id = inscripcion.curso.id
    nombre_alumno = inscripcion.alumno.get_full_name() or inscripcion.alumno.username
    titulo_curso = inscripcion.curso.titulo
    
    registrar_log(request, "Desmatriculación", f"Desmatriculó a '{nombre_alumno}' del curso '{titulo_curso}'")
    inscripcion.delete()
    messages.success(request, f"Se desmatriculó al alumno '{nombre_alumno}' del curso '{titulo_curso}'.")
    return redirect('admin_curso_alumnos', curso_id=curso_id)


# ----------------------------------------------------
# SECCIÓN ADMINISTRACIÓN: CARGA MASIVA, REPORTES Y LOGS
# ----------------------------------------------------
@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_carga_masiva_usuarios(request):
    if request.method == 'POST' and request.FILES.get('archivo_csv'):
        archivo = request.FILES['archivo_csv']
        
        if not archivo.name.lower().endswith('.csv'):
            messages.error(request, "El archivo seleccionado debe tener extensión .csv")
            return redirect('admin_carga_masiva_usuarios')

        try:
            # 1. Leer y decodificar con tolerancia a formatos de Excel
            raw_data = archivo.read()
            try:
                contenido = raw_data.decode('utf-8-sig')
            except UnicodeDecodeError:
                contenido = raw_data.decode('latin-1')

            # 2. Detectar si Excel usó comas (,) o punto y coma (;)
            primera_linea = contenido.split('\n')[0] if contenido else ''
            delimitador = ';' if ';' in primera_linea and ',' not in primera_linea else ','

            lector = csv.DictReader(io.StringIO(contenido), delimiter=delimitador)
            
            # Limpiar posibles espacios accidentales en los encabezados
            if lector.fieldnames:
                lector.fieldnames = [nombre.strip().lower() for nombre in lector.fieldnames if nombre]

            creados = 0
            omitidos = 0

            for fila in lector:
                # Obtener valores limpiando espacios
                username = fila.get('username', '').strip()
                email = fila.get('email', '').strip()
                first_name = fila.get('first_name', '').strip()
                last_name = fila.get('last_name', '').strip()
                password = fila.get('password', '').strip()
                rol = fila.get('rol', 'Alumnos').strip() or 'Alumnos'

                if username and password and not User.objects.filter(username__iexact=username).exists():
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        first_name=first_name,
                        last_name=last_name
                    )
                    
                    # Asignar permisos si es Administrador
                    if rol.lower() in ['administrador', 'admin']:
                        user.is_staff = True
                        user.save()
                    
                    # Asignar grupo
                    grupo, _ = Group.objects.get_or_create(name=rol)
                    user.groups.add(grupo)
                    creados += 1
                else:
                    omitidos += 1

            registrar_log(request, "Carga Masiva", f"Creó {creados} usuarios vía CSV ({omitidos} omitidos)")
            messages.success(request, f"Carga masiva completada: {creados} usuarios creados con éxito ({omitidos} omitidos o duplicados).")
            return redirect('admin_dashboard')

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo CSV: {str(e)}")
            return redirect('admin_carga_masiva_usuarios')

    return render(request, 'admin_carga_masiva.html')

@login_required
def descargar_plantilla_usuarios(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("No tienes permiso para realizar esta acción.")

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="plantilla_carga_usuarios.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['username', 'first_name', 'last_name', 'email', 'password', 'rol'])
    writer.writerow(['jperez', 'Juan', 'Perez Garcia', 'jperez@galeno.pe', 'Temporal123*', 'Alumnos'])
    writer.writerow(['mrodriguez', 'Maria', 'Rodriguez Soto', 'mrodriguez@galeno.pe', 'Temporal123*', 'Docentes'])
    
    return response

@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def exportar_usuarios_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios_galeno.csv"'
    response.write('\ufeff'.encode('utf8'))

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Usuario', 'Nombres', 'Apellidos', 'Correo', 'Rol', 'Fecha de Registro'])

    for u in User.objects.all().prefetch_related('groups').order_by('last_name', 'first_name'):
        if u.is_superuser or u.is_staff:
            rol = 'Administrador'
        elif u.groups.filter(name='Docentes').exists():
            rol = 'Docente'
        else:
            rol = 'Alumno'

        writer.writerow([
            u.username,
            u.first_name,
            u.last_name,
            u.email or 'Sin correo',
            rol,
            u.date_joined.strftime('%d/%m/%Y %H:%M')
        ])

    registrar_log(request, "Exportación de Datos", "Descargó el listado general de usuarios en CSV")
    return response


@login_required
@user_passes_test(es_administrador, login_url='/cuentas/login/')
def admin_logs_actividad(request):
    total_alumnos = User.objects.filter(groups__name='Alumnos').count()
    total_docentes = User.objects.filter(groups__name='Docentes').count()
    total_cursos = Curso.objects.count()

    query = request.GET.get('q', '').strip()
    logs_qs = LogActividad.objects.all().select_related('usuario').order_by('-fecha')

    if query:
        logs_qs = logs_qs.filter(
            Q(accion__icontains=query) |
            Q(detalles__icontains=query) |
            Q(usuario__username__icontains=query)
        )

    paginator = Paginator(logs_qs, 20)
    page_number = request.GET.get('page')
    logs = paginator.get_page(page_number)

    context = {
        'total_alumnos': total_alumnos,
        'total_docentes': total_docentes,
        'total_cursos': total_cursos,
        'logs': logs,
        'query': query,
    }
    return render(request, 'admin_logs.html', context)


# ----------------------------------------------------
# SECCIÓN DOCENTE
# ----------------------------------------------------
@login_required
def panel_docente(request):
    if not es_docente_valido(request.user):
        return HttpResponseForbidden("Acceso exclusivo para docentes.")

    if request.user.is_superuser or request.user.is_staff:
        cursos = Curso.objects.prefetch_related('docentes', 'inscripciones', 'materiales').all()
    else:
        cursos = Curso.objects.filter(docentes=request.user).prefetch_related('docentes', 'inscripciones', 'materiales').distinct()

    return render(request, 'panel_docente.html', {'cursos': cursos})


@login_required
def subir_material(request, curso_id):
    if not es_docente_valido(request.user):
        return HttpResponseForbidden("No tienes permiso para realizar esta acción.")
    
    curso = get_object_or_404(Curso, id=curso_id)

    if not (request.user.is_staff or request.user.is_superuser) and not curso.docentes.filter(id=request.user.id).exists():
        return HttpResponseForbidden("No estás asignado como docente en este curso.")
    
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        semana = request.POST.get('semana', 1)
        archivo = request.FILES.get('archivo')
        enlace = request.POST.get('enlace') or request.POST.get('enlace_web')
        
        if titulo:
            Material.objects.create(
                curso=curso,
                titulo=titulo,
                semana=semana,
                archivo=archivo,
                enlace_web=enlace  # <-- AQUÍ: cambiar enlace por enlace_web
            )
            registrar_log(request, "Subida de Material", f"Subió '{titulo}' (Semana {semana}) al curso '{curso.titulo}'")
            messages.success(request, f"Material '{titulo}' publicado exitosamente en la Semana {semana}.")
            return redirect('detalle_curso', curso_id=curso.id)
            
    return render(request, 'subir_material.html', {'curso': curso})


@login_required
def eliminar_material(request, material_id):
    material = get_object_or_404(Material, id=material_id)
    curso = material.curso
    es_autorizado = request.user.is_staff or request.user.is_superuser or curso.docentes.filter(id=request.user.id).exists()

    if not es_autorizado:
        return HttpResponseForbidden("No tienes permiso para eliminar este material.")

    titulo_mat = material.titulo
    material.delete()
    registrar_log(request, "Eliminación de Material", f"Eliminó el material '{titulo_mat}' del curso '{curso.titulo}'")
    messages.success(request, f"Material '{titulo_mat}' eliminado correctamente.")
    return redirect('detalle_curso', curso_id=curso.id)


@login_required
def docente_calificar_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)

    es_docente = curso.docentes.filter(id=request.user.id).exists()
    es_admin = request.user.is_superuser or request.user.groups.filter(name='Administrador').exists()
    
    if not (es_docente or es_admin):
        messages.error(request, "No tienes permisos para calificar en este curso.")
        return redirect('panel_docente')

    inscripciones = Inscripcion.objects.filter(curso=curso).select_related('alumno')

    def parsear_nota(valor):
        if valor is not None and str(valor).strip() != '':
            try:
                return float(valor)
            except ValueError:
                return None
        return None

    if request.method == 'POST':
        for insc in inscripciones:
            alumno_id = str(insc.alumno.id)
            n1 = request.POST.get(f'nota1_{alumno_id}')
            n2 = request.POST.get(f'nota2_{alumno_id}')
            n3 = request.POST.get(f'nota3_{alumno_id}')

            calificacion, _ = Calificacion.objects.get_or_create(curso=curso, alumno=insc.alumno)
            calificacion.nota1 = parsear_nota(n1)
            calificacion.nota2 = parsear_nota(n2)
            calificacion.nota3 = parsear_nota(n3)
            calificacion.save()

        registrar_log(request, "Registro de Calificaciones", f"Actualizó notas del curso '{curso.titulo}'")
        messages.success(request, "Las calificaciones se guardaron correctamente.")
        return redirect('docente_calificar_curso', curso_id=curso.id)

    calificaciones_dict = {c.alumno_id: c for c in Calificacion.objects.filter(curso=curso)}
    filas_calificaciones = []
    for insc in inscripciones:
        calif = calificaciones_dict.get(insc.alumno.id)
        filas_calificaciones.append({
            'alumno': insc.alumno,
            'nota1': calif.nota1 if calif and calif.nota1 is not None else '',
            'nota2': calif.nota2 if calif and calif.nota2 is not None else '',
            'nota3': calif.nota3 if calif and calif.nota3 is not None else '',
            'promedio': calif.promedio if calif and hasattr(calif, 'promedio') else None
        })

    return render(request, 'docente_calificar.html', {
        'curso': curso,
        'filas_calificaciones': filas_calificaciones
    })


# ----------------------------------------------------
# VISTAS GENERALES DE ALUMNO Y CURSOS
# ----------------------------------------------------
@login_required
def detalle_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)
    es_docente_curso = curso.docentes.filter(id=request.user.id).exists()
    esta_matriculado = Inscripcion.objects.filter(alumno=request.user, curso=curso).exists()
    
    if not (esta_matriculado or es_docente_curso or request.user.is_staff or request.user.is_superuser):
        return HttpResponseForbidden("No tienes permiso para ver este curso.")
    
    periodo = curso.periodo or PeriodoAcademico.objects.filter(activo=True).first()

    cronograma = []
    if periodo:
        cronograma = periodo.obtener_cronograma_semanas()
    else:
        cronograma = [{'numero': i, 'inicio': None, 'fin': None, 'etiqueta': f"Semana {i}"} for i in range(1, 11)]

    # Consultas de contenidos
    materiales = list(curso.materiales.all().order_by('semana', '-fecha_subida'))
    examenes = list(curso.examenes.filter(activo=True).prefetch_related('preguntas'))
    
    bloques_semanas = []
    for sem in cronograma:
        mats_semana = [m for m in materiales if m.semana == sem['numero']]
        exams_semana = [e for e in examenes if e.semana == sem['numero']]
        
        bloques_semanas.append({
            'info': sem,
            'materiales': mats_semana,
            'examenes': exams_semana,
            'total_materiales': len(mats_semana),
            'total_recursos': len(mats_semana) + len(exams_semana)
        })

    notas = Calificacion.objects.filter(alumno=request.user, curso=curso)
    
    context = {
        'curso': curso,
        'materiales': materiales,
        'bloques_semanas': bloques_semanas,
        'cronograma': cronograma,
        'notas': notas,
        'periodo': periodo,
        'es_docente_curso': es_docente_curso or request.user.is_staff or request.user.is_superuser,
    }

    return render(request, 'detalle_curso.html', context)

@login_required
def mis_notas(request):
    periodos = PeriodoAcademico.objects.all().order_by('-fecha_inicio')
    periodo_id = request.GET.get('periodo')
    
    # 1. Determinar el período a consultar (por URL o el que esté marcado como Activo)
    periodo_actual = None
    if periodo_id:
        periodo_actual = PeriodoAcademico.objects.filter(id=periodo_id).first()
    if not periodo_actual:
        periodo_actual = PeriodoAcademico.objects.filter(activo=True).first() or periodos.first()

    # 2. Filtro estricto por el período seleccionado
    if periodo_actual:
        inscripciones = Inscripcion.objects.filter(
            alumno=request.user,
            curso__periodo=periodo_actual
        ).select_related('curso')
    else:
        inscripciones = Inscripcion.objects.none()

    cursos_alumno = [ins.curso for ins in inscripciones]

    context = {
        'periodos': periodos,
        'periodo_actual': periodo_actual,
        'cursos_alumno': cursos_alumno,
    }
    return render(request, 'notas.html', context)
@login_required
def mi_perfil(request):
    return render(request, 'perfil.html')

@login_required
def gestionar_temporada(request):
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('admin_dashboard')

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'crear':
            nombre = request.POST.get('nombre', '').strip()
            codigo = request.POST.get('codigo', '').strip()
            fecha_inicio = request.POST.get('fecha_inicio')
            fecha_fin = request.POST.get('fecha_fin')
            activar_inmediato = request.POST.get('activo') == 'on'

            # Validar si el código ya existe
            if PeriodoAcademico.objects.filter(codigo=codigo).exists():
                messages.error(request, f"Ya existe un período registrado con el código '{codigo}'. Usa otro código.")
                return redirect('admin_dashboard')

            if activar_inmediato:
                PeriodoAcademico.objects.update(activo=False)

            PeriodoAcademico.objects.create(
                nombre=nombre,
                codigo=codigo,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                activo=activar_inmediato
            )
            messages.success(request, f"Temporada '{nombre}' creada exitosamente.")

        elif accion == 'culminar':
            periodo_id = request.POST.get('periodo_id')
            periodo = PeriodoAcademico.objects.filter(id=periodo_id).first()
            if periodo:
                periodo.activo = False
                periodo.save()
                messages.warning(request, f"La temporada '{periodo.nombre}' ha sido culminada. El ciclo quedó cerrado.")

    return redirect('admin_dashboard')

@login_required
def banco_preguntas_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)

    if not (request.user.is_staff or getattr(request.user, 'rol', None) == 'docente' or request.user.is_superuser):
        messages.error(request, "No tienes permisos para gestionar el banco de preguntas.")
        return redirect('detalle_curso', curso_id=curso.id)

    # Obtenemos los exámenes existentes creados para este curso
    examenes_curso = Examen.objects.filter(curso=curso)

    if request.method == 'POST':
        form = PreguntaForm(request.POST, curso=curso)
        if form.is_valid():
            pregunta = form.save()
            correcta = form.cleaned_data['opcion_correcta']

            Opcion.objects.create(pregunta=pregunta, texto=form.cleaned_data['opcion_1'], es_correcta=(correcta == '1'))
            Opcion.objects.create(pregunta=pregunta, texto=form.cleaned_data['opcion_2'], es_correcta=(correcta == '2'))
            Opcion.objects.create(pregunta=pregunta, texto=form.cleaned_data['opcion_3'], es_correcta=(correcta == '3'))
            Opcion.objects.create(pregunta=pregunta, texto=form.cleaned_data['opcion_4'], es_correcta=(correcta == '4'))

            messages.success(request, "Pregunta agregada con éxito al banco.")
            return redirect('banco_preguntas_curso', curso_id=curso.id)
    else:
        form = PreguntaForm(curso=curso)

    preguntas = Pregunta.objects.filter(examen__curso=curso).prefetch_related('opciones').order_by('-id')

    context = {
        'curso': curso,
        'form': form,
        'preguntas': preguntas,
        'examenes_curso': examenes_curso,
    }
    return render(request, 'banco_preguntas.html', context)

@login_required
def rendir_examen(request, examen_id):
    examen = get_object_or_404(Examen, id=examen_id)

    # 1. Validación de disponibilidad para alumnos
    es_docente = (
        examen.curso.docentes.filter(id=request.user.id).exists()
        or request.user.is_staff
        or request.user.is_superuser
    )
    if not es_docente and not examen.esta_disponible:
        messages.error(request, f"Acceso restringido: {examen.estado_texto}.")
        return redirect('detalle_curso', curso_id=examen.curso.id)

    # 2. Obtener preguntas aleatorias (Pool)
    queryset_preguntas = examen.preguntas.order_by('?')
    if examen.cantidad_preguntas_aleatorias > 0:
        preguntas = list(queryset_preguntas[:examen.cantidad_preguntas_aleatorias])
    else:
        preguntas = list(queryset_preguntas)

    # 3. Barajar alternativas
    for pregunta in preguntas:
        pregunta.opciones_aleatorias = list(pregunta.opciones.order_by('?'))

    # 4. Calificación sobre las preguntas que realmente le tocaron
    if request.method == 'POST':
        puntaje_total = 0.0
        # Evaluamos solo sobre las preguntas que se enviaron en el intento
        for pregunta in examen.preguntas.all():
            opcion_seleccionada_id = request.POST.get(f'pregunta_{pregunta.id}')
            if opcion_seleccionada_id:
                try:
                    opcion = Opcion.objects.get(id=opcion_seleccionada_id, pregunta=pregunta)
                    if opcion.es_correcta:
                        puntaje_total += float(pregunta.puntaje)
                except Opcion.DoesNotExist:
                    pass

        IntentoExamen.objects.create(
            alumno=request.user,
            examen=examen,
            nota=puntaje_total,
            completado=True,
            fecha_fin=timezone.now()
        )
        messages.success(request, f"Examen finalizado. Tu puntaje obtenido es: {puntaje_total} puntos.")
        return redirect('detalle_curso', curso_id=examen.curso.id)

    return render(request, 'rendir_examen.html', {'examen': examen, 'preguntas': preguntas})

@login_required
def descargar_plantilla_preguntas(request):
    """Genera directamente un archivo binario .xlsx con columnas independientes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BancoPreguntas"

    # Encabezados: cada uno es una celda independiente
    headers = [
        "Enunciado de la Pregunta",
        "Alternativa A",
        "Alternativa B",
        "Alternativa C",
        "Alternativa D",
        "Respuesta Correcta (A, B, C o D)",
        "Puntaje",
        "Explicación Clínica"
    ]
    ws.append(headers)

    # Estilos de encabezado
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Fila de ejemplo con datos en cada celda
    ejemplo = [
        "¿Cuál es el agente causal más frecuente de la infección del tracto urinario?",
        "Escherichia coli",
        "Staphylococcus aureus",
        "Klebsiella pneumoniae",
        "Pseudomonas aeruginosa",
        "A",
        2.0,
        "E. coli representa más del 80% de los casos comunitarios."
    ]
    ws.append(ejemplo)

    for col_idx in range(1, len(ejemplo) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    # Anchos fijos por columna
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 45

    # Guardar en memoria binaria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_preguntas.xlsx"'
    return response

@login_required
def importar_preguntas_curso(request, curso_id):
    """Lee el archivo .xlsx subido por el profesor y crea las preguntas."""
    curso = get_object_or_404(Curso, id=curso_id)

    if not (request.user.is_staff or getattr(request.user, 'rol', None) == 'docente' or request.user.is_superuser):
        messages.error(request, "No tienes permisos para esta acción.")
        return redirect('detalle_curso', curso_id=curso.id)

    if request.method == 'POST' and request.FILES.get('archivo_preguntas'):
        examen_id = request.POST.get('examen_id')
        archivo = request.FILES['archivo_preguntas']

        if not examen_id:
            messages.error(request, "Debes seleccionar una evaluación de destino.")
            return redirect('banco_preguntas_curso', curso_id=curso.id)

        examen = get_object_or_404(Examen, id=examen_id, curso=curso)

        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active

            creadas = 0
            # Iterar desde la fila 2 para saltar la cabecera
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue

                enunciado = str(row[0]).strip()
                op_a = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                op_b = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                op_c = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                op_d = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                correcta = str(row[5]).strip().upper() if len(row) > 5 and row[5] is not None else "A"

                try:
                    puntaje = float(row[6]) if len(row) > 6 and row[6] is not None else 1.0
                except (ValueError, TypeError):
                    puntaje = 1.0

                explicacion = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""

                if not enunciado or not op_a:
                    continue

                # Guardar en Base de Datos
                pregunta = Pregunta.objects.create(
                    examen=examen,
                    enunciado=enunciado,
                    explicacion=explicacion,
                    puntaje=puntaje
                )

                Opcion.objects.create(pregunta=pregunta, texto=op_a, es_correcta=(correcta == 'A'))
                Opcion.objects.create(pregunta=pregunta, texto=op_b, es_correcta=(correcta == 'B'))
                Opcion.objects.create(pregunta=pregunta, texto=op_c, es_correcta=(correcta == 'C'))
                Opcion.objects.create(pregunta=pregunta, texto=op_d, es_correcta=(correcta == 'D'))

                creadas += 1

            messages.success(request, f"Se importaron con éxito {creadas} preguntas a la evaluación '{examen.titulo}'.")
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo Excel: {str(e)}")
    else:
        messages.error(request, "Por favor adjunta un archivo Excel válido (.xlsx).")

    return redirect('banco_preguntas_curso', curso_id=curso.id)

@login_required
def descargar_plantilla_preguntas(request):
    """Genera una plantilla Excel (.xlsx) con celdas, colores y columnas tabuladas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Preguntas"

    # Encabezados de columna
    headers = [
        "Enunciado de la Pregunta",
        "Alternativa A",
        "Alternativa B",
        "Alternativa C",
        "Alternativa D",
        "Respuesta Correcta (A, B, C o D)",
        "Puntaje",
        "Explicación Clínica (Opcional)"
    ]
    ws.append(headers)

    # Estilos de encabezado
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Fila de ejemplo
    ws.append([
        "¿Cuál es el agente causal más frecuente de la infección del tracto urinario?",
        "Escherichia coli",
        "Staphylococcus aureus",
        "Klebsiella pneumoniae",
        "Pseudomonas aeruginosa",
        "A",
        2.0,
        "E. coli representa más del 80% de las ITUs comunitarias."
    ])

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_num)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    # Ajuste de ancho de columnas
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_preguntas_galeno.xlsx"'
    wb.save(response)
    return response


@login_required
def importar_preguntas_curso(request, curso_id):
    """Lee el archivo .xlsx cargado y registra preguntas y opciones en la BD."""
    curso = get_object_or_404(Curso, id=curso_id)

    if not (request.user.is_staff or getattr(request.user, 'rol', None) == 'docente' or request.user.is_superuser):
        messages.error(request, "No tienes permisos para esta acción.")
        return redirect('detalle_curso', curso_id=curso.id)

    if request.method == 'POST' and request.FILES.get('archivo_preguntas'):
        examen_id = request.POST.get('examen_id')
        archivo = request.FILES['archivo_preguntas']

        if not examen_id:
            messages.error(request, "Debes seleccionar una evaluación de destino.")
            return redirect('banco_preguntas_curso', curso_id=curso.id)

        examen = get_object_or_404(Examen, id=examen_id, curso=curso)

        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active

            creadas = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue

                enunciado = str(row[0]).strip()
                op_a = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                op_b = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                op_c = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                op_d = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                correcta = str(row[5]).strip().upper() if len(row) > 5 and row[5] is not None else "A"

                try:
                    puntaje = float(row[6]) if len(row) > 6 and row[6] is not None else 1.0
                except (ValueError, TypeError):
                    puntaje = 1.0

                explicacion = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""

                if not enunciado or not op_a:
                    continue

                pregunta = Pregunta.objects.create(
                    examen=examen,
                    enunciado=enunciado,
                    explicacion=explicacion,
                    puntaje=puntaje
                )

                Opcion.objects.create(pregunta=pregunta, texto=op_a, es_correcta=(correcta == 'A'))
                Opcion.objects.create(pregunta=pregunta, texto=op_b, es_correcta=(correcta == 'B'))
                Opcion.objects.create(pregunta=pregunta, texto=op_c, es_correcta=(correcta == 'C'))
                Opcion.objects.create(pregunta=pregunta, texto=op_d, es_correcta=(correcta == 'D'))

                creadas += 1

            messages.success(request, f"Se importaron con éxito {creadas} preguntas a '{examen.titulo}'.")
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo Excel: {str(e)}")
    else:
        messages.error(request, "Por favor adjunta un archivo Excel válido (.xlsx).")

    return redirect('banco_preguntas_curso', curso_id=curso.id)

@login_required
def crear_examen_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)

    if not (request.user.is_staff or getattr(request.user, 'rol', None) == 'docente' or request.user.is_superuser):
        messages.error(request, "No tienes permisos para programar evaluaciones.")
        return redirect('detalle_curso', curso_id=curso.id)

    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        semana = request.POST.get('semana', 1)
        duracion = request.POST.get('duracion_minutos', 60)
        descripcion = request.POST.get('descripcion', '')

        if titulo:
            Examen.objects.create(
                curso=curso,
                titulo=titulo,
                semana=int(semana),
                duracion_minutos=int(duracion),
                descripcion=descripcion,
                activo=True
            )
            messages.success(request, f"Evaluación '{titulo}' creada con éxito en la Semana {semana}.")
        else:
            messages.error(request, "Debes ingresar un título para la evaluación.")

    return redirect('detalle_curso', curso_id=curso.id)

@login_required
def toggle_examen(request, examen_id):
    """Permite al docente pausar o habilitar el examen para los alumnos con un solo clic."""
    examen = get_object_or_404(Examen, id=examen_id)
    es_docente = (
        examen.curso.docentes.filter(id=request.user.id).exists()
        or request.user.is_staff
        or request.user.is_superuser
    )
    if not es_docente:
        return HttpResponseForbidden("No tienes permisos para realizar esta acción.")

    examen.activo = not examen.activo
    examen.save()
    estado = "habilitado" if examen.activo else "pausado"
    messages.success(request, f"Evaluación '{examen.titulo}' {estado} correctamente.")
    return redirect('detalle_curso', curso_id=examen.curso.id)


@login_required
def eliminar_examen(request, examen_id):
    """Permite al docente eliminar una evaluación programada."""
    examen = get_object_or_404(Examen, id=examen_id)
    curso_id = examen.curso.id
    es_docente = (
        examen.curso.docentes.filter(id=request.user.id).exists()
        or request.user.is_staff
        or request.user.is_superuser
    )
    if not es_docente:
        return HttpResponseForbidden("No tienes permisos para realizar esta acción.")

    titulo = examen.titulo
    examen.delete()
    messages.success(request, f"Evaluación '{titulo}' eliminada del sistema.")
    return redirect('detalle_curso', curso_id=curso_id)

@login_required
def eliminar_pregunta(request, pregunta_id):
    """Permite al docente o administrador eliminar una pregunta específica de una evaluación."""
    pregunta = get_object_or_404(Pregunta, id=pregunta_id)
    curso = pregunta.examen.curso

    # Validar que el usuario sea docente asignado al curso, staff o superusuario
    es_docente = (
        curso.docentes.filter(id=request.user.id).exists()
        or request.user.is_staff
        or request.user.is_superuser
    )
    if not es_docente:
        return HttpResponseForbidden("No tienes permisos para eliminar preguntas de este curso.")

    # Eliminar la pregunta (sus opciones asociadas se eliminan en cascada)
    pregunta.delete()
    messages.success(request, "Pregunta eliminada correctamente del banco de la evaluación.")
    return redirect('banco_preguntas_curso', curso_id=curso.id)

@login_required
def rendir_examen(request, examen_id):
    examen = get_object_or_404(Examen, id=examen_id)

    es_docente = (
        examen.curso.docentes.filter(id=request.user.id).exists()
        or request.user.is_staff
        or request.user.is_superuser
    )

    # 1. Validaciones para Alumnos (Disponibilidad e Intentos)
    if not es_docente:
        if not examen.esta_disponible:
            messages.error(request, f"Acceso restringido: {examen.estado_texto}.")
            return redirect('detalle_curso', curso_id=examen.curso.id)

        intentos_hechos = IntentoExamen.objects.filter(
            alumno=request.user, 
            examen=examen, 
            completado=True
        ).count()

        if intentos_hechos >= examen.intentos_permitidos:
            messages.warning(request, f"Has alcanzado el límite máximo de intentos permitidos ({examen.intentos_permitidos}).")
            return redirect('revision_examen', examen_id=examen.id)

    # 2. Obtener Preguntas (Pool aleatorio)
    queryset_preguntas = examen.preguntas.order_by('?')
    if getattr(examen, 'cantidad_preguntas_aleatorias', 0) > 0:
        preguntas = list(queryset_preguntas[:examen.cantidad_preguntas_aleatorias])
    else:
        preguntas = list(queryset_preguntas)

    for p in preguntas:
        p.opciones_aleatorias = list(p.opciones.order_by('?'))

    # 3. Procesar Envío (Manual o por agotamiento del Timer)
    if request.method == 'POST':
        puntaje_total = 0.0

        intento = IntentoExamen.objects.create(
            alumno=request.user,
            examen=examen,
            completado=True,
            fecha_fin=timezone.now()
        )

        for pregunta in examen.preguntas.all():
            opcion_id = request.POST.get(f'pregunta_{pregunta.id}')
            if opcion_id:
                try:
                    opcion = Opcion.objects.get(id=opcion_id, pregunta=pregunta)
                    es_correcta = opcion.es_correcta
                    if es_correcta:
                        puntaje_total += float(pregunta.puntaje)
                    
                    # Registrar respuesta para revisión diferida
                    RespuestaEstudiante.objects.create(
                        intento=intento,
                        pregunta=pregunta,
                        opcion_seleccionada=opcion,
                        es_correcta=es_correcta
                    )
                except Opcion.DoesNotExist:
                    pass

        intento.nota = puntaje_total
        intento.save()

        messages.success(request, f"Evaluación finalizada. Tu nota es: {puntaje_total} puntos.")
        return redirect('revision_examen', examen_id=examen.id)

    context = {
        'examen': examen,
        'preguntas': preguntas,
        'tiempo_segundos': examen.duracion_minutos * 60,
    }
    return render(request, 'rendir_examen.html', context)


@login_required
def revision_examen(request, examen_id):
    """Permite ver el resultado y la retroalimentación médica/teórica."""
    examen = get_object_or_404(Examen, id=examen_id)
    
    # Obtener el último intento realizado por el estudiante
    ultimo_intento = IntentoExamen.objects.filter(
        alumno=request.user, 
        examen=examen, 
        completado=True
    ).order_by('-fecha_fin').first()

    if not ultimo_intento and not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "No has realizado ningún intento en esta evaluación.")
        return redirect('detalle_curso', curso_id=examen.curso.id)

    respuestas = ultimo_intento.respuestas.select_related('pregunta', 'opcion_seleccionada').all() if ultimo_intento else []

    context = {
        'examen': examen,
        'intento': ultimo_intento,
        'respuestas': respuestas,
        'puede_ver_solucionario': examen.revision_disponible or request.user.is_staff,
    }
    return render(request, 'revision_examen.html', context)